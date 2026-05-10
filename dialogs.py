"""
Dialog windows for the Swiss Chess Tournament Manager.
"""
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout, QGridLayout,
    QLineEdit, QSpinBox, QComboBox, QPushButton, QLabel, QGroupBox,
    QListWidget, QListWidgetItem, QDialogButtonBox, QCheckBox,
    QFileDialog, QMessageBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView
)
from datetime import datetime
from PyQt6.QtCore import Qt
from models import (
    Player, Team, Tournament, TournamentType, Title, TieBreakType,
    TSF_2025_DEFAULT_TIEBREAKS, MatchResult
)
from database import get_all_tournaments_info, delete_tournament_from_db
from fide_integration import search_fide_xml, fide_to_player
from PyQt6.QtGui import QIcon
import os
import sys

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


class TournamentDialog(QDialog):
    """Create or edit tournament settings."""

    def __init__(self, tournament: Tournament = None, parent=None):
        super().__init__(parent)
        self.tournament = tournament or Tournament()
        self.setWindowTitle("Turnuva Ayarları")
        self.setMinimumWidth(550)
        self._init_ui()
        self._load_data()

    def _init_ui(self):
        layout = QVBoxLayout(self)

        # Basic info
        info_group = QGroupBox("Turnuva Bilgileri")
        form = QFormLayout(info_group)

        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("Turnuva adı...")
        form.addRow("Turnuva Adı:", self.name_edit)

        self.type_combo = QComboBox()
        self.type_combo.addItem("İsviçre Sistemi", TournamentType.SWISS)
        self.type_combo.addItem("Round-Robin (Berger)", TournamentType.ROUND_ROBIN)
        self.type_combo.addItem("Takım İsviçre", TournamentType.TEAM_SWISS)
        form.addRow("Turnuva Türü:", self.type_combo)

        self.date_edit = QLineEdit()
        self.date_edit.setPlaceholderText("GG.AA.YYYY")
        form.addRow("Tarih:", self.date_edit)

        self.location_edit = QLineEdit()
        form.addRow("Konum:", self.location_edit)

        self.arbiter_edit = QLineEdit()
        form.addRow("Hakem:", self.arbiter_edit)

        self.rounds_spin = QSpinBox()
        self.rounds_spin.setRange(1, 23)
        self.rounds_spin.setValue(7)
        form.addRow("Tur Sayısı:", self.rounds_spin)

        self.double_rr_check = QCheckBox("Çift Round-Robin")
        form.addRow("", self.double_rr_check)

        layout.addWidget(info_group)

        # Tiebreak settings
        tb_group = QGroupBox("Eşitlik Bozma Sıralaması")
        tb_layout = QVBoxLayout(tb_group)

        self.tb_list = QListWidget()
        self.tb_list.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        tb_layout.addWidget(self.tb_list)

        tb_btn_layout = QHBoxLayout()
        self.tb_add_btn = QPushButton("Ekle")
        self.tb_add_btn.clicked.connect(self._add_tiebreak)
        self.tb_remove_btn = QPushButton("Kaldır")
        self.tb_remove_btn.clicked.connect(self._remove_tiebreak)
        self.tb_remove_btn.setProperty("danger", True)
        self.tb_reset_btn = QPushButton("TSF 2025")
        self.tb_reset_btn.clicked.connect(self._reset_tiebreaks)
        self.tb_reset_btn.setProperty("secondary", True)

        tb_btn_layout.addWidget(self.tb_add_btn)
        tb_btn_layout.addWidget(self.tb_remove_btn)
        tb_btn_layout.addWidget(self.tb_reset_btn)
        tb_btn_layout.addStretch()
        tb_layout.addLayout(tb_btn_layout)

        layout.addWidget(tb_group)

        # Buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._save_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _load_data(self):
        self.name_edit.setText(self.tournament.name)
        idx = self.type_combo.findData(self.tournament.tournament_type)
        if idx >= 0:
            self.type_combo.setCurrentIndex(idx)
        self.date_edit.setText(self.tournament.date)
        self.location_edit.setText(self.tournament.location)
        self.arbiter_edit.setText(self.tournament.arbiter)
        self.rounds_spin.setValue(self.tournament.num_rounds)
        self.double_rr_check.setChecked(self.tournament.is_double_round_robin)

        self.tb_list.clear()
        for tb in self.tournament.tiebreak_order:
            item = QListWidgetItem(f"{tb.value} — {tb.display_name}")
            item.setData(Qt.ItemDataRole.UserRole, tb)
            self.tb_list.addItem(item)

    def _add_tiebreak(self):
        current_tbs = set()
        for i in range(self.tb_list.count()):
            current_tbs.add(self.tb_list.item(i).data(Qt.ItemDataRole.UserRole))

        available = [tb for tb in TieBreakType if tb not in current_tbs]
        if not available:
            QMessageBox.information(self, "Bilgi", "Tüm tie-break sistemleri zaten ekli.")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Tie-Break Ekle")
        layout = QVBoxLayout(dlg)
        combo = QComboBox()
        for tb in available:
            combo.addItem(f"{tb.value} — {tb.display_name}", tb)
        layout.addWidget(combo)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        layout.addWidget(btns)

        if dlg.exec() == QDialog.DialogCode.Accepted:
            tb = combo.currentData()
            item = QListWidgetItem(f"{tb.value} — {tb.display_name}")
            item.setData(Qt.ItemDataRole.UserRole, tb)
            self.tb_list.addItem(item)

    def _remove_tiebreak(self):
        row = self.tb_list.currentRow()
        if row >= 0:
            self.tb_list.takeItem(row)

    def _reset_tiebreaks(self):
        self.tb_list.clear()
        for tb in TSF_2025_DEFAULT_TIEBREAKS:
            item = QListWidgetItem(f"{tb.value} — {tb.display_name}")
            item.setData(Qt.ItemDataRole.UserRole, tb)
            self.tb_list.addItem(item)

    def _save_and_accept(self):
        self.tournament.name = self.name_edit.text() or "Yeni Turnuva"
        self.tournament.tournament_type = self.type_combo.currentData()
        self.tournament.date = self.date_edit.text()
        self.tournament.location = self.location_edit.text()
        self.tournament.arbiter = self.arbiter_edit.text()
        self.tournament.num_rounds = self.rounds_spin.value()
        self.tournament.is_double_round_robin = self.double_rr_check.isChecked()

        self.tournament.tiebreak_order = []
        for i in range(self.tb_list.count()):
            tb = self.tb_list.item(i).data(Qt.ItemDataRole.UserRole)
            self.tournament.tiebreak_order.append(tb)

        self.accept()

    def get_tournament(self) -> Tournament:
        return self.tournament


class PlayerDialog(QDialog):
    """Add or edit a player."""

    def __init__(self, player: Player = None, teams: list[Team] = None, parent=None):
        super().__init__(parent)
        self.player = player or Player()
        self.teams = teams or []
        self.setWindowTitle("Oyuncu Düzenle" if player else "Oyuncu Ekle")
        self.setMinimumWidth(420)
        self._init_ui()
        self._load_data()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        form = QFormLayout()

        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("Ad")
        form.addRow("Ad:", self.name_edit)

        self.surname_edit = QLineEdit()
        self.surname_edit.setPlaceholderText("Soyad")
        form.addRow("Soyad:", self.surname_edit)

        self.rating_spin = QSpinBox()
        self.rating_spin.setRange(0, 3000)
        self.rating_spin.setToolTip("Eğer UKD veya ELO girilirse otomatik hesaplanır.")
        form.addRow("Rating:", self.rating_spin)

        self.ukd_spin = QSpinBox()
        self.ukd_spin.setRange(0, 3000)
        form.addRow("UKD:", self.ukd_spin)

        self.elo_spin = QSpinBox()
        self.elo_spin.setRange(0, 3500)
        form.addRow("ELO:", self.elo_spin)

        self.registered_check = QCheckBox("Kayıt Tamamlandı")
        form.addRow("", self.registered_check)

        self.title_combo = QComboBox()
        for t in Title:
            label = t.value if t != Title.NONE else "(Yok)"
            self.title_combo.addItem(label, t)
        form.addRow("Ünvan:", self.title_combo)

        self.club_edit = QLineEdit()
        form.addRow("Kulüp:", self.club_edit)

        self.birth_spin = QSpinBox()
        self.birth_spin.setRange(0, 2026)
        self.birth_spin.setSpecialValueText("—")
        form.addRow("Doğum Yılı:", self.birth_spin)

        self.fide_edit = QLineEdit()
        self.fide_edit.setPlaceholderText("FIDE ID")
        form.addRow("FIDE ID:", self.fide_edit)

        self.team_combo = QComboBox()
        self.team_combo.addItem("(Takım yok)", "")
        for team in self.teams:
            self.team_combo.addItem(team.name, team.id)
        form.addRow("Takım:", self.team_combo)

        layout.addLayout(form)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._save_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _load_data(self):
        self.name_edit.setText(self.player.name)
        self.surname_edit.setText(self.player.surname)
        self.rating_spin.setValue(self.player.rating)
        self.ukd_spin.setValue(self.player.ukd)
        self.elo_spin.setValue(self.player.elo)
        self.registered_check.setChecked(self.player.is_registered)
        idx = self.title_combo.findData(self.player.title)
        if idx >= 0:
            self.title_combo.setCurrentIndex(idx)
        self.club_edit.setText(self.player.club)
        self.birth_spin.setValue(self.player.birth_year)
        self.fide_edit.setText(self.player.fide_id)
        team_idx = self.team_combo.findData(self.player.team_id)
        if team_idx >= 0:
            self.team_combo.setCurrentIndex(team_idx)

    def _save_and_accept(self):
        self.player.name = self.name_edit.text()
        self.player.surname = self.surname_edit.text()
        self.player.rating = self.rating_spin.value()
        self.player.ukd = self.ukd_spin.value()
        self.player.elo = self.elo_spin.value()
        self.player.is_registered = self.registered_check.isChecked()
        self.player.update_rating() # Updates rating if ukd/elo are larger and rating wasn't manually set larger
        self.player.title = self.title_combo.currentData()
        self.player.club = self.club_edit.text()
        self.player.birth_year = self.birth_spin.value()
        self.player.fide_id = self.fide_edit.text()
        self.player.team_id = self.team_combo.currentData() or ""
        self.accept()

    def get_player(self) -> Player:
        return self.player


class TeamDialog(QDialog):
    """Add or edit a team."""

    def __init__(self, team: Team = None, parent=None):
        super().__init__(parent)
        self.team = team or Team()
        self.setWindowTitle("Takım Düzenle" if team else "Takım Ekle")
        self.setMinimumWidth(350)
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        form = QFormLayout()

        self.name_edit = QLineEdit()
        self.name_edit.setText(self.team.name)
        self.name_edit.setPlaceholderText("Takım adı")
        form.addRow("Takım Adı:", self.name_edit)
        layout.addLayout(form)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._save_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _save_and_accept(self):
        self.team.name = self.name_edit.text() or "Takım"
        self.accept()

    def get_team(self) -> Team:
        return self.team


class FideSearchDialog(QDialog):
    """Search FIDE XML rating list and import players."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("FIDE Rating Listesinden Oyuncu Ara")
        self.setMinimumSize(650, 450)
        self.selected_players: list[Player] = []
        self.fide_filepath = ""
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)

        file_layout = QHBoxLayout()
        self.file_label = QLabel("FIDE XML dosyası seçilmedi")
        file_btn = QPushButton("XML Seç")
        file_btn.setProperty("secondary", True)
        file_btn.clicked.connect(self._select_file)
        file_layout.addWidget(self.file_label, 1)
        file_layout.addWidget(file_btn)
        layout.addLayout(file_layout)

        search_layout = QHBoxLayout()
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("İsim veya FIDE ID ile ara...")
        self.search_edit.returnPressed.connect(self._search)
        search_btn = QPushButton("Ara")
        search_btn.clicked.connect(self._search)
        search_layout.addWidget(self.search_edit, 1)
        search_layout.addWidget(search_btn)
        layout.addLayout(search_layout)

        self.results_table = QTableWidget()
        self.results_table.setColumnCount(5)
        self.results_table.setHorizontalHeaderLabels(["İsim", "FIDE ID", "Rating", "Ünvan", "Ülke"])
        self.results_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.results_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.results_table.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        layout.addWidget(self.results_table)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._import_selected)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _select_file(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self, "FIDE XML Rating Listesi", "", "XML Files (*.xml);;All Files (*)"
        )
        if filepath:
            self.fide_filepath = filepath
            self.file_label.setText(filepath.split("/")[-1].split("\\")[-1])

    def _search(self):
        if not self.fide_filepath:
            QMessageBox.warning(self, "Uyarı", "Önce bir FIDE XML dosyası seçin.")
            return

        query = self.search_edit.text().strip()
        if not query:
            return

        results = search_fide_xml(self.fide_filepath, query)
        self.results_table.setRowCount(len(results))

        for row, data in enumerate(results):
            self.results_table.setItem(row, 0, QTableWidgetItem(data.get('name', '')))
            self.results_table.setItem(row, 1, QTableWidgetItem(data.get('fide_id', '')))
            self.results_table.setItem(row, 2, QTableWidgetItem(str(data.get('rating', 0))))
            self.results_table.setItem(row, 3, QTableWidgetItem(data.get('title', '')))
            self.results_table.setItem(row, 4, QTableWidgetItem(data.get('country', '')))

        self._search_results = results

    def _import_selected(self):
        selected_rows = set()
        for item in self.results_table.selectedItems():
            selected_rows.add(item.row())

        if hasattr(self, '_search_results'):
            for row in selected_rows:
                if row < len(self._search_results):
                    player = fide_to_player(self._search_results[row])
                    self.selected_players.append(player)

        self.accept()

    def get_selected_players(self) -> list[Player]:
        return self.selected_players


class PlayerListImportDialog(QDialog):
    """Import players from CSV, Excel (.xlsx), or tab-separated text files."""

    # Column name aliases for auto-detection (Turkish + English)
    _ALIASES = {
        "name": ["ad", "adı", "name", "first name", "firstname", "isim"],
        "surname": ["soyad", "soyadı", "soyadi", "soyisim", "surname", "last name", "lastname"],
        "rating": ["rating", "rtg", "puan", "derece"],
        "ukd": ["ukd"],
        "elo": ["elo"],
        "title": ["unvan", "ünvan", "title", "tit"],
        "club": ["kulüp", "kulup", "club", "team", "takım", "takim"],
        "fide_id": ["fide id", "fide_id", "fideid", "fide no", "fide"],
        "birth_year": ["doğum yılı", "dogum yili", "birth year", "birthyear", "yıl", "yil", "born"],
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Oyuncu Listesi İçe Aktar")
        self.setMinimumSize(800, 550)
        self.imported_players: list[Player] = []
        self._raw_data: list[dict] = []
        self._col_mapping: dict[str, str] = {}
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)

        # File selection
        file_layout = QHBoxLayout()
        self.file_label = QLabel("Dosya seçilmedi")
        file_btn = QPushButton("📂 Dosya Seç")
        file_btn.clicked.connect(self._select_file)
        file_layout.addWidget(self.file_label, 1)
        file_layout.addWidget(file_btn)
        layout.addLayout(file_layout)

        # Info label
        self.info_label = QLabel("Desteklenen formatlar: CSV, Excel (.xlsx), TXT/TSV (sekme ayrılmış), NAT (.nat)")
        self.info_label.setStyleSheet("color: #666; font-size: 11px;")
        layout.addWidget(self.info_label)

        # Column mapping group
        map_group = QGroupBox("Sütun Eşleştirme")
        map_layout = QFormLayout(map_group)

        self.map_combos: dict[str, QComboBox] = {}
        field_labels = {
            "name": "Ad:", "surname": "Soyad:", "rating": "Rating:",
            "ukd": "UKD:", "elo": "ELO:",
            "title": "Ünvan:", "club": "Kulüp:", "fide_id": "FIDE ID:",
            "birth_year": "Doğum Yılı:",
        }
        for key, label in field_labels.items():
            combo = QComboBox()
            combo.addItem("(Atla)", "")
            self.map_combos[key] = combo
            map_layout.addRow(label, combo)

        layout.addWidget(map_group)

        # Preview table
        preview_group = QGroupBox("Önizleme")
        preview_layout = QVBoxLayout(preview_group)
        self.preview_table = QTableWidget()
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.preview_table.verticalHeader().setVisible(False)
        preview_layout.addWidget(self.preview_table)
        self.preview_count = QLabel("0 oyuncu bulundu")
        preview_layout.addWidget(self.preview_count)
        layout.addWidget(preview_group)

        # Buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._do_import)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _select_file(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self, "Oyuncu Listesi Seç", "",
            "Tüm Desteklenen (*.csv *.xlsx *.txt *.tsv *.nat);;CSV (*.csv);;Excel (*.xlsx);;Text (*.txt *.tsv);;NAT (*.nat)"
        )
        if not filepath:
            return

        basename = filepath.replace("\\", "/").split("/")[-1]
        self.file_label.setText(basename)

        try:
            if filepath.lower().endswith(".xlsx"):
                self._raw_data = self._read_excel(filepath)
            else:
                self._raw_data = self._read_csv(filepath)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Dosya okunamadı:\n{e}")
            return

        if not self._raw_data:
            QMessageBox.warning(self, "Uyarı", "Dosyada veri bulunamadı.")
            return

        self._setup_column_mapping()
        self._update_preview()

    def _read_csv(self, filepath: str) -> list[dict]:
        """Read CSV or TSV file. Auto-detects delimiter."""
        import csv
        rows = []
        with open(filepath, "r", encoding="utf-8-sig") as f:
            sample = f.read(2048)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(f, dialect=dialect)
            for row in reader:
                rows.append(row)
        return rows

    def _read_excel(self, filepath: str) -> list[dict]:
        """Read Excel .xlsx file using openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            QMessageBox.critical(self, "Hata", "openpyxl yüklü değil.\npip install openpyxl")
            return []

        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        headers = next(rows_iter, None)
        if not headers:
            return []
        headers = [str(h).strip() if h else f"Col{i}" for i, h in enumerate(headers)]

        data = []
        for row in rows_iter:
            if all(v is None for v in row):
                continue
            entry = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    entry[headers[i]] = str(val).strip() if val is not None else ""
            data.append(entry)
        wb.close()
        return data

    def _setup_column_mapping(self):
        """Populate column mapping combos and auto-detect matches."""
        if not self._raw_data:
            return

        file_columns = list(self._raw_data[0].keys())

        for key, combo in self.map_combos.items():
            combo.clear()
            combo.addItem("(Atla)", "")
            for col in file_columns:
                combo.addItem(col, col)

            # Auto-detect
            aliases = self._ALIASES.get(key, [])
            matched = False
            for col in file_columns:
                col_lower = col.replace("İ", "i").replace("I", "ı").lower().strip()
                if col_lower in aliases:
                    idx = combo.findData(col)
                    if idx >= 0:
                        combo.setCurrentIndex(idx)
                        matched = True
                        break

            # Fallback: try partial match
            if not matched:
                for col in file_columns:
                    col_lower = col.replace("İ", "i").replace("I", "ı").lower().strip()
                    if key == "name" and "soy" in col_lower:
                        continue
                    for alias in aliases:
                        if alias in col_lower or col_lower in alias:
                            idx = combo.findData(col)
                            if idx >= 0:
                                combo.setCurrentIndex(idx)
                                matched = True
                                break
                    if matched:
                        break

        # Connect change signals for live preview
        for combo in self.map_combos.values():
            combo.currentIndexChanged.connect(self._update_preview)

    def _get_mapping(self) -> dict[str, str]:
        """Get current field->column mapping."""
        mapping = {}
        for key, combo in self.map_combos.items():
            col = combo.currentData()
            if col:
                mapping[key] = col
        return mapping

    def _parse_players(self) -> list[Player]:
        """Parse raw data into Player objects using current mapping."""
        mapping = self._get_mapping()
        players = []

        for row in self._raw_data:
            name = row.get(mapping.get("name", ""), "").strip()
            surname = row.get(mapping.get("surname", ""), "").strip()

            # Skip empty rows
            if not name and not surname:
                continue

            rating = 0
            raw_rating = row.get(mapping.get("rating", ""), "")
            try:
                rating = int(str(raw_rating).strip().replace(".", "").replace(",", ""))
            except (ValueError, TypeError):
                pass
                
            ukd = 0
            raw_ukd = row.get(mapping.get("ukd", ""), "")
            try:
                ukd = int(str(raw_ukd).strip().replace(".", "").replace(",", ""))
            except (ValueError, TypeError):
                pass
                
            elo = 0
            raw_elo = row.get(mapping.get("elo", ""), "")
            try:
                elo = int(str(raw_elo).strip().replace(".", "").replace(",", ""))
            except (ValueError, TypeError):
                pass

            title_str = row.get(mapping.get("title", ""), "").strip().upper()
            title = Title.NONE
            for t in Title:
                if t.value == title_str:
                    title = t
                    break

            club = row.get(mapping.get("club", ""), "").strip()
            fide_id = row.get(mapping.get("fide_id", ""), "").strip()

            birth_year = 0
            raw_by = row.get(mapping.get("birth_year", ""), "")
            try:
                birth_year = int(str(raw_by).strip())
            except (ValueError, TypeError):
                pass

            player = Player(
                name=name, surname=surname, rating=rating,
                ukd=ukd, elo=elo,
                title=title, club=club, fide_id=fide_id,
                birth_year=birth_year,
            )
            player.update_rating()
            players.append(player)

        return players

    def _update_preview(self):
        """Refresh preview table based on current column mapping."""
        players = self._parse_players()
        headers = ["Ad", "Soyad", "Rating", "UKD", "ELO", "Ünvan", "Kulüp", "FIDE ID"]
        self.preview_table.setColumnCount(len(headers))
        self.preview_table.setHorizontalHeaderLabels(headers)
        self.preview_table.setRowCount(min(len(players), 50))

        for row, p in enumerate(players[:50]):
            self.preview_table.setItem(row, 0, QTableWidgetItem(p.name))
            self.preview_table.setItem(row, 1, QTableWidgetItem(p.surname))
            self.preview_table.setItem(row, 2, QTableWidgetItem(str(p.rating) if p.rating > 0 else ""))
            self.preview_table.setItem(row, 3, QTableWidgetItem(str(p.ukd) if p.ukd > 0 else ""))
            self.preview_table.setItem(row, 4, QTableWidgetItem(str(p.elo) if p.elo > 0 else ""))
            self.preview_table.setItem(row, 5, QTableWidgetItem(p.display_title))
            self.preview_table.setItem(row, 6, QTableWidgetItem(p.club))
            self.preview_table.setItem(row, 7, QTableWidgetItem(p.fide_id))

        self.preview_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.preview_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        suffix = " (ilk 50 gösteriliyor)" if len(players) > 50 else ""
        self.preview_count.setText(f"{len(players)} oyuncu bulundu{suffix}")

    def _do_import(self):
        self.imported_players = self._parse_players()
        if not self.imported_players:
            QMessageBox.warning(self, "Uyarı", "İçe aktarılacak oyuncu bulunamadı.")
            return
        self.accept()

    def get_players(self) -> list[Player]:
        return self.imported_players


class PlayerDetailsDialog(QDialog):
    """View player details and match history."""

    def __init__(self, player: Player, tournament: Tournament, parent=None):
        super().__init__(parent)
        self.player = player
        self.tournament = tournament
        self.setWindowTitle(f"Oyuncu Detayları: {player.full_name}")
        self.setMinimumSize(600, 400)
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)

        # Player Info Group
        info_group = QGroupBox("Kişisel Bilgiler")
        info_layout = QGridLayout(info_group)

        # Labels
        info_layout.addWidget(QLabel("<b>İsim Soyisim:</b>"), 0, 0)
        info_layout.addWidget(QLabel(self.player.full_name), 0, 1)

        info_layout.addWidget(QLabel("<b>Rating:</b>"), 1, 0)
        info_layout.addWidget(QLabel(str(self.player.rating) if self.player.rating > 0 else "—"), 1, 1)

        info_layout.addWidget(QLabel("<b>Ünvan:</b>"), 0, 2)
        info_layout.addWidget(QLabel(self.player.display_title if self.player.display_title else "—"), 0, 3)

        info_layout.addWidget(QLabel("<b>FIDE ID:</b>"), 1, 2)
        info_layout.addWidget(QLabel(self.player.fide_id if self.player.fide_id else "—"), 1, 3)

        info_layout.addWidget(QLabel("<b>Kulüp:</b>"), 2, 0)
        info_layout.addWidget(QLabel(self.player.club if self.player.club else "—"), 2, 1)

        layout.addWidget(info_group)

        # Match History Group
        history_group = QGroupBox("Maç Geçmişi")
        history_layout = QVBoxLayout(history_group)

        self.history_table = QTableWidget()
        self.history_table.setAlternatingRowColors(True)
        self.history_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.history_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.history_table.verticalHeader().setVisible(False)

        columns = ["Tur", "Masa", "Renk", "Rakip", "Rakip Rtg", "Sonuç"]
        self.history_table.setColumnCount(len(columns))
        self.history_table.setHorizontalHeaderLabels(columns)

        self._populate_history()
        history_layout.addWidget(self.history_table)
        layout.addWidget(history_group)

        # Close button
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _populate_history(self):
        matches_data = []
        for rnd in self.tournament.rounds:
            for m in rnd.matches:
                if m.white_id == self.player.id or m.black_id == self.player.id:
                    matches_data.append((rnd.round_number, m))
                    break

        self.history_table.setRowCount(len(matches_data))
        for row, (rnd_num, match) in enumerate(matches_data):
            is_white = match.white_id == self.player.id

            # Tur
            item = QTableWidgetItem(str(rnd_num))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.history_table.setItem(row, 0, item)

            # Masa
            item = QTableWidgetItem(str(match.board))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.history_table.setItem(row, 1, item)

            # Renk
            color_text = "Beyaz" if is_white else "Siyah"
            if match.is_bye:
                color_text = "—"
            item = QTableWidgetItem(color_text)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.history_table.setItem(row, 2, item)

            # Rakip & Rakip Rtg
            if match.is_bye:
                opp_name = "BYE"
                opp_rtg = "—"
            else:
                opp_id = match.black_id if is_white else match.white_id
                opp = self.tournament.get_player_by_id(opp_id)
                opp_name = opp.full_name if opp else "Bilinmiyor"
                opp_rtg = str(opp.rating) if opp and opp.rating > 0 else "—"

            self.history_table.setItem(row, 3, QTableWidgetItem(opp_name))

            item = QTableWidgetItem(opp_rtg)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.history_table.setItem(row, 4, item)

            # Sonuç
            res_str = "—"
            from models import MatchResult
            if match.is_bye:
                if match.result == MatchResult.HALF_POINT_BYE: res_str = "½ BYE"
                elif match.result == MatchResult.ZERO_POINT_BYE: res_str = "0 BYE"
                else: res_str = "1 BYE"
            else:
                if match.result == MatchResult.WHITE_WIN: res_str = "1 - 0" if is_white else "0 - 1"
                elif match.result == MatchResult.BLACK_WIN: res_str = "0 - 1" if is_white else "1 - 0"
                elif match.result == MatchResult.DRAW: res_str = "½ - ½"
                elif match.result == MatchResult.WHITE_FORFEIT: res_str = "+ (f)" if is_white else "- (f)"
                elif match.result == MatchResult.BLACK_FORFEIT: res_str = "- (f)" if is_white else "+ (f)"
                elif match.result == MatchResult.DOUBLE_FORFEIT: res_str = "- (f)"

            item = QTableWidgetItem(res_str)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            font = item.font()
            font.setBold(True)
            item.setFont(font)
            self.history_table.setItem(row, 5, item)

        self.history_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        for i in [0, 1, 2, 4, 5]:
            self.history_table.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)

class DashboardDialog(QDialog):
    """Startup dashboard to select or create tournaments from the SQLite database."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("PAU SCTM - Turnuva Yöneticisi")
        self.setWindowIcon(QIcon(resource_path("resources/logo.png")))
        self.resize(600, 400)
        self.selected_tournament_id = None
        self._init_ui()
        self._load_data()

    def _init_ui(self):
        from PyQt6.QtWidgets import QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView, QMessageBox
        layout = QVBoxLayout(self)

        # Header
        header = QLabel("Kayıtlı Turnuvalar")
        font = header.font()
        font.setPointSize(14)
        font.setBold(True)
        header.setFont(font)
        layout.addWidget(header)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Turnuva Adı", "Tarih", "Son Değişiklik"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.doubleClicked.connect(self._on_open)
        layout.addWidget(self.table)

        # Buttons
        btn_layout = QHBoxLayout()
        
        self.btn_new = QPushButton("✨ Yeni Turnuva")
        self.btn_new.clicked.connect(self._on_new)
        btn_layout.addWidget(self.btn_new)

        btn_layout.addStretch()

        self.btn_delete = QPushButton("🗑️ Sil")
        self.btn_delete.clicked.connect(self._on_delete)
        btn_layout.addWidget(self.btn_delete)

        self.btn_open = QPushButton("📂 Turnuvayı Aç")
        self.btn_open.setDefault(True)
        self.btn_open.clicked.connect(self._on_open)
        btn_layout.addWidget(self.btn_open)

        layout.addLayout(btn_layout)

    def _load_data(self):
        from database import get_all_tournaments_info
        from PyQt6.QtWidgets import QTableWidgetItem
        from PyQt6.QtCore import Qt
        
        self.tournaments = get_all_tournaments_info()
        self.table.setRowCount(len(self.tournaments))
        for row, t in enumerate(self.tournaments):
            self.table.setItem(row, 0, QTableWidgetItem(t["name"]))
            self.table.setItem(row, 1, QTableWidgetItem(t["date"]))
            # Format datetime
            dt_str = t["last_modified"]
            try:
                dt_obj = datetime.fromisoformat(dt_str)
                display_dt = dt_obj.strftime("%Y-%m-%d %H:%M")
            except:
                display_dt = dt_str
            self.table.setItem(row, 2, QTableWidgetItem(display_dt))
            
            # Store ID in the first item
            self.table.item(row, 0).setData(Qt.ItemDataRole.UserRole, t["id"])

    def _on_new(self):
        self.selected_tournament_id = None
        self.accept()

    def _on_open(self):
        from PyQt6.QtWidgets import QMessageBox
        from PyQt6.QtCore import Qt
        
        row = self.table.currentRow()
        if row >= 0:
            item = self.table.item(row, 0)
            self.selected_tournament_id = item.data(Qt.ItemDataRole.UserRole)
            self.accept()
        else:
            QMessageBox.warning(self, "Uyarı", "Lütfen açmak için bir turnuva seçin.")

    def _on_delete(self):
        from PyQt6.QtWidgets import QMessageBox
        from PyQt6.QtCore import Qt
        from database import delete_tournament_from_db
        
        row = self.table.currentRow()
        if row < 0:
            return
            
        item = self.table.item(row, 0)
        t_id = item.data(Qt.ItemDataRole.UserRole)
        name = item.text()
        
        reply = QMessageBox.question(
            self, "Onay", 
            f"'{name}' turnuvasını silmek istediğinize emin misiniz?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            delete_tournament_from_db(t_id)
            self._load_data()


class UnpairedPlayersDialog(QDialog):
    """Dialog to show players who will not be paired in the upcoming round."""

    def __init__(self, unpaired_players: list[Player], missed_rounds_map: dict[str, list[str]], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Eşlendirilmeyen Oyuncular")
        self.setMinimumWidth(450)
        self.setMinimumHeight(300)
        
        layout = QVBoxLayout(self)
        
        label = QLabel("Aşağıdaki oyuncular bu turda eşlendirilmeyecek.\nDevam etmek istiyor musunuz?")
        layout.addWidget(label)
        
        self.list_widget = QListWidget()
        for p in unpaired_players:
            past_missed = len(missed_rounds_map.get(p.id, []))
            
            if past_missed > 0:
                text = f"{p.full_name} (Geçmiş {past_missed} Tura Katılmadı)"
            else:
                text = p.full_name
                
            item = QListWidgetItem(text)
            
            if past_missed >= 2:
                item.setForeground(Qt.GlobalColor.red)
                font = item.font()
                font.setStrikeOut(True)
                item.setFont(font)
                
            self.list_widget.addItem(item)
            
        layout.addWidget(self.list_widget)
        
        btn_layout = QHBoxLayout()
        btn_yes = QPushButton("Evet, Eşlendir")
        btn_yes.clicked.connect(self.accept)
        btn_no = QPushButton("İptal")
        btn_no.clicked.connect(self.reject)
        
        btn_layout.addWidget(btn_yes)
        btn_layout.addWidget(btn_no)
        layout.addLayout(btn_layout)
