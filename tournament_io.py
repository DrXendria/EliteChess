"""
Tournament save/load (JSON .swt) and export (HTML, Excel).
"""
import json
import uuid
import os
from typing import Optional
from models import (
    Tournament, Player, Team, Match, Round, MatchResult,
    TournamentType, Title, TieBreakType
)
from tiebreak import calculate_all_standings


def save_tournament(tournament: Tournament, filepath: str) -> bool:
    """Save tournament to JSON .swt file."""
    try:
        data = _tournament_to_dict(tournament)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"Save error: {e}")
        return False


def load_tournament(filepath: str) -> Optional[Tournament]:
    """Load tournament from JSON .swt file."""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return _dict_to_tournament(data)
    except Exception as e:
        print(f"Load error: {e}")
        return None


def _tournament_to_dict(t: Tournament) -> dict:
    return {
        "id": t.id,
        "name": t.name,
        "tournament_type": t.tournament_type.value,
        "date": t.date,
        "location": t.location,
        "arbiter": t.arbiter,
        "num_rounds": t.num_rounds,
        "current_round": t.current_round,
        "is_double_round_robin": t.is_double_round_robin,
        "tiebreak_order": [tb.value for tb in t.tiebreak_order],
        "players": [_player_to_dict(p) for p in t.players],
        "teams": [_team_to_dict(tm) for tm in t.teams],
        "rounds": [_round_to_dict(r) for r in t.rounds],
    }


def _player_to_dict(p: Player) -> dict:
    return {
        "id": p.id, "name": p.name, "surname": p.surname,
        "rating": p.rating, "ukd": p.ukd, "elo": p.elo,
        "title": p.title.value,
        "club": p.club, "birth_year": p.birth_year,
        "fide_id": p.fide_id, "team_id": p.team_id,
        "is_active": p.is_active,
        "is_registered": p.is_registered,
        "withdrawn_after_round": p.withdrawn_after_round,
    }


def _team_to_dict(t: Team) -> dict:
    return {"id": t.id, "name": t.name, "player_ids": t.player_ids}


def _round_to_dict(r: Round) -> dict:
    return {
        "round_number": r.round_number,
        "is_completed": r.is_completed,
        "matches": [_match_to_dict(m) for m in r.matches],
    }


def _match_to_dict(m: Match) -> dict:
    return {
        "board": m.board, "white_id": m.white_id,
        "black_id": m.black_id, "result": m.result.value,
    }


def _dict_to_tournament(d: dict) -> Tournament:
    t = Tournament(
        id=d.get("id", str(uuid.uuid4())),
        name=d.get("name", "Yeni Turnuva"),
        tournament_type=TournamentType(d.get("tournament_type", "swiss")),
    )
    t.date = d.get("date", "")
    t.location = d.get("location", "")
    t.arbiter = d.get("arbiter", "")
    t.num_rounds = d.get("num_rounds", 7)
    t.current_round = d.get("current_round", 0)
    t.is_double_round_robin = d.get("is_double_round_robin", False)
    t.tiebreak_order = [TieBreakType(v) for v in d.get("tiebreak_order", [])]
    t.players = [_dict_to_player(p) for p in d.get("players", [])]
    t.teams = [_dict_to_team(tm) for tm in d.get("teams", [])]
    t.rounds = [_dict_to_round(r) for r in d.get("rounds", [])]
    return t


def _dict_to_player(d: dict) -> Player:
    return Player(
        id=d["id"], name=d.get("name", ""), surname=d.get("surname", ""),
        rating=d.get("rating", 0), ukd=d.get("ukd", 0), elo=d.get("elo", 0),
        title=Title(d.get("title", "")),
        club=d.get("club", ""), birth_year=d.get("birth_year", 0),
        fide_id=d.get("fide_id", ""), team_id=d.get("team_id", ""),
        is_active=d.get("is_active", True),
        is_registered=d.get("is_registered", False),
        withdrawn_after_round=d.get("withdrawn_after_round", 0),
    )


def _dict_to_team(d: dict) -> Team:
    return Team(id=d["id"], name=d.get("name", ""), player_ids=d.get("player_ids", []))


def _dict_to_round(d: dict) -> Round:
    return Round(
        round_number=d["round_number"],
        is_completed=d.get("is_completed", False),
        matches=[_dict_to_match(m) for m in d.get("matches", [])],
    )


def _dict_to_match(d: dict) -> Match:
    return Match(
        board=d["board"], white_id=d["white_id"],
        black_id=d["black_id"], result=MatchResult(d.get("result", "none")),
    )


# ─── HTML EXPORT ─────────────────────────────────────────────────────────────

def export_html(tournament: Tournament, filepath: str) -> bool:
    """Export tournament results as HTML (chess-results.com compatible format)."""
    try:
        standings = calculate_all_standings(tournament)
        html = _build_html(tournament, standings)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html)
        return True
    except Exception as e:
        print(f"HTML export error: {e}")
        return False


def _build_html(tournament: Tournament, standings: list[dict]) -> str:
    tb_headers = "".join(
        f"<th>{tb.value}</th>" for tb in tournament.tiebreak_order
    )

    rows = ""
    for entry in standings:
        p = entry["player"]
        tb_cells = "".join(
            f"<td>{entry['tiebreaks'].get(tb, 0):.1f}</td>"
            for tb in tournament.tiebreak_order
        )
        rows += f"""<tr>
            <td>{entry['rank']}</td>
            <td>{p.display_title}</td>
            <td>{p.full_name}</td>
            <td>{p.rating}</td>
            <td>{p.club}</td>
            <td><b>{entry['score']:.1f}</b></td>
            {tb_cells}
        </tr>\n"""

    round_tables = ""
    for rnd in tournament.rounds:
        round_rows = ""
        for m in rnd.matches:
            wp = tournament.get_player_by_id(m.white_id)
            bp = tournament.get_player_by_id(m.black_id) if m.black_id != "BYE" else None
            w_name = wp.full_name if wp else m.white_id
            b_name = bp.full_name if bp else "BYE"
            result_str = m.result.value if m.result.is_decided else "-"
            round_rows += f"""<tr>
                <td>{m.board}</td>
                <td>{w_name}</td>
                <td>{result_str}</td>
                <td>{b_name}</td>
            </tr>\n"""

        round_tables += f"""
        <h3>Tur {rnd.round_number}</h3>
        <table class="round-table">
            <tr><th>Masa</th><th>Beyaz</th><th>Sonuç</th><th>Siyah</th></tr>
            {round_rows}
        </table>\n"""

    return f"""<!DOCTYPE html>
<html lang="tr">
<head>
    <meta charset="UTF-8">
    <title>{tournament.name} - Sonuçlar</title>
    <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
        h1 {{ color: #1a1a2e; }}
        h2 {{ color: #16213e; border-bottom: 2px solid #0f3460; padding-bottom: 5px; }}
        h3 {{ color: #16213e; }}
        .info {{ margin-bottom: 20px; color: #555; }}
        table {{ border-collapse: collapse; width: 100%; margin-bottom: 20px; background: white; }}
        th {{ background: #1a1a2e; color: white; padding: 8px 12px; text-align: left; }}
        td {{ padding: 6px 12px; border-bottom: 1px solid #ddd; }}
        tr:nth-child(even) {{ background: #f8f9fa; }}
        tr:hover {{ background: #e8f0fe; }}
        .round-table {{ width: auto; max-width: 600px; }}
        b {{ color: #0f3460; }}
    </style>
</head>
<body>
    <h1>{tournament.name}</h1>
    <div class="info">
        <p><b>Tarih:</b> {tournament.date} | <b>Konum:</b> {tournament.location} | <b>Hakem:</b> {tournament.arbiter}</p>
        <p><b>Sistem:</b> {tournament.tournament_type.value.replace('_', ' ').title()} | <b>Turlar:</b> {tournament.current_round}/{tournament.num_rounds}</p>
    </div>

    <h2>Sıralama</h2>
    <table>
        <tr>
            <th>Sıra</th><th>Ünvan</th><th>İsim</th><th>Rating</th><th>Kulüp</th><th>Puan</th>
            {tb_headers}
        </tr>
        {rows}
    </table>

    {round_tables}

    <footer style="margin-top: 30px; color: #999; font-size: 12px;">
        Swiss Chess Tournament Manager ile oluşturulmuştur.
    </footer>
</body>
</html>"""


# ─── EXCEL EXPORT ────────────────────────────────────────────────────────────

def export_excel(tournament: Tournament, filepath: str) -> bool:
    """Export tournament results as Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # Standings sheet
        ws = wb.active
        ws.title = "Sıralama"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        standings = calculate_all_standings(tournament)

        headers = ["Sıra", "Ünvan", "İsim", "Rating", "Kulüp", "Puan"]
        headers += [tb.value for tb in tournament.tiebreak_order]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, entry in enumerate(standings, 2):
            p = entry["player"]
            values = [
                entry["rank"], p.display_title, p.full_name,
                p.rating, p.club, entry["score"]
            ]
            values += [entry["tiebreaks"].get(tb, 0) for tb in tournament.tiebreak_order]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                if col >= 6:
                    cell.alignment = Alignment(horizontal='center')

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Rounds sheets
        for rnd in tournament.rounds:
            ws_r = wb.create_sheet(f"Tur {rnd.round_number}")
            r_headers = ["Masa", "Beyaz", "Sonuç", "Siyah"]
            for col, h in enumerate(r_headers, 1):
                cell = ws_r.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            for row_idx, m in enumerate(rnd.matches, 2):
                wp = tournament.get_player_by_id(m.white_id)
                bp = tournament.get_player_by_id(m.black_id) if m.black_id != "BYE" else None
                values = [
                    m.board,
                    wp.full_name if wp else m.white_id,
                    m.result.value if m.result.is_decided else "-",
                    bp.full_name if bp else "BYE",
                ]
                for col, val in enumerate(values, 1):
                    cell = ws_r.cell(row=row_idx, column=col, value=val)
                    cell.border = thin_border

        wb.save(filepath)
        return True
    except Exception as e:
        print(f"Excel export error: {e}")
        return False
